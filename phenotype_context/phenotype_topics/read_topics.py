'''
Reading phenotype topics from files.
'''

from openpyxl import load_workbook
from phenotype_context.settings import phenoTopicsLocation

phenoTable, phenoSheet = phenoTopicsLocation


def parsePhenotypeTable(file, sheet, verbose=False):
    '''
    Create mapping of pheno id (ordinal number in the order of appearance) to pheno data
    map of corpusId/centroid -> list of (word, weight) pairs
    '''
    wb = load_workbook(file)
    ws = wb[sheet]

    def isEmptyRow(r, emptyCells=5):
        '''
        :param emptyCells: number of cells, from rows start, that have to be empty
                for the row to be judged empty
        '''
        for i in range(emptyCells):
            if r[i].value != None: return False
        return True

    def isHeaderRow(r):
        return r[0].value == None and r[1].value != None

    def isCorpusRow(r):
        return r[0].value != None and r[1].value != None

    def readRow(r):
        '''
        Read and parse row of cells with word,weight string.
        :return: list of (word string, value float) pairs
        '''
        res = []; corpusId = None
        for i, c in enumerate(r):
            val = c.value
            if i == 0:
                if val is None: corpusId='CENTROID'
                else: corpusId = str(c.value)
                continue
            if val is None: break # empty cell, end
            val = val.strip()
            sval = val.split(',')
            word, weight = sval[0], float(sval[1])
            res.append((word, weight))
        return corpusId, res

    phenoid2parse = {}; phenoId = 0; numEmpty = 0
    corpus2words = {}; nonadded = False
    for i, r in enumerate(ws):
        if i == 0: continue # skip header row
        if isEmptyRow(r):
            numEmpty += 1
            if numEmpty > 1: break # end of spreadsheet
            # end of phenotype, add parsed data 2 map
            phenoid2parse[phenoId] = corpus2words
            #print i, corpus2words
            corpus2words = {}; phenoId += 1; nonadded = False
        else:
            numEmpty = 0
            cid, words = readRow(r)
            corpus2words[cid] = words
            nonadded = True
            if verbose:
                print cid+':', ' '.join('%s, %.3f' % ww for ww in words)
    if nonadded: # end of table, add last phenotype parse
        phenoid2parse[phenoId] = corpus2words
    return phenoid2parse

def readPhenotypesFromTable(file, sheet, verbose=False):
    wb = load_workbook(file)
    ws = wb[sheet]

    def isEmptyRow(r, emptyCells=5):
        '''
        :param emptyCells: number of cells, from rows start, that have to be empty
                for the row to be judged empty
        '''
        for i in range(emptyCells):
            if r[i].value != None: return False
        return True

    def isHeaderRow(r):
        return r[0].value == None and r[1].value != None

    def readRow(r, startAt=1):
        '''
        Read and parse row of cells with word,weight string.
        :return: list of (word string, value float) pairs
        '''
        res = []
        for i, c in enumerate(r):
            if i == 0: continue # skip first cell
            val = c.value
            if val is None: break
            val = val.strip()
            sval = val.split(',')
            word, weight = sval[0], float(sval[1])
            res.append((word, weight))
        return res

    phenos = []
    numEmpty = 0
    for i, r in enumerate(ws):
        if i == 0: continue # skip header row
        if isEmptyRow(r):
            numEmpty += 1
            if numEmpty == 5: break # 5 empty rows => end of spreadsheet
        else:
            numEmpty = 0
            if isHeaderRow(r):
                wws = readRow(r)
                phenos.append(wws)
                if verbose:
                    print ' '.join('%s, %.3f' % ww for ww in wws)
    if verbose: print len(phenos)
    return phenos

def loadTable(): return readPhenotypesFromTable(phenoTable, phenoSheet)
def loadParse(): return parsePhenotypeTable(phenoTable, phenoSheet)

def validate():
    ''' Assert equivalence of old and new parses '''
    old = readPhenotypesFromTable(phenoTable, phenoSheet)
    new = parsePhenotypeTable(phenoTable, phenoSheet)
    assert len(old) == len(new)
    print len(new)
    for i, ww in enumerate(old):
        if i not in new or new[i]['CENTROID'] != ww:
            print i, ww

if __name__ == '__main__':
    #print len(readPhenotypesFromTable(phenoTable, phenoSheet))
    #print len(parsePhenotypeTable(phenoTable, phenoSheet))
    #for ph in parsePhenotypeTable(phenoTable, phenoSheet, True): print ph
    #for ph in readPhenotypesFromTable(phenoTable, phenoSheet, True): print ph
    validate()